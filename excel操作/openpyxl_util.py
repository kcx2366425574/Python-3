# -*- encoding : utf-8 -*-
"""
@File       : openpyxl_util.py
@Time       :2021/5/6 17:11
@Author     :kuang congxian
@Contact    :kuangcx@inspur.com
@Description : null
"""
import ast

import openpyxl

# 创建工作簿
f = openpyxl.Workbook()

# 创建sheet
sheet1 = f.create_sheet()


def json_to_xlsx(json_path):
    with open(json_path, "r") as file:
        data = ast.literal_eval(file.read())

        # 往表格写入标题
        i = 1
        for key in data[0].keys():
            sheet1.cell(row=1, column=i).value = key
            i += 1

        if type(data) == list:
            i = 2
            for d in data:
                j = 1
                for key, value in d.items():
                    sheet1.cell(row=i, column=j).value = value
                    j += 1
                i += 1

        f.save("data.xlsx")  # 保存文件


if __name__ == '__main__':
    json_to_xlsx("data.json")
